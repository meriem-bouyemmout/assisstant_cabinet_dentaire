from tkinter import *
from tkinter import ttk
import customtkinter as ctk
from PIL import Image,ImageTk
import tkinter.messagebox as mb
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pathlib
import sqlite3
from tkinter import filedialog

class Assistante:
    def __init__(self,mast):
        self.master = mast
        self.master.title("Assistante Cabinet Dentaire")
        self.width = self.master.winfo_screenwidth()
        self.height = self.master.winfo_screenheight()
        self.master.geometry("{w}x{h}+0+0".format(w=self.width,h=self.height))
        self.master.state("zoomed")
        # ctk.set_appearance_mode("system")  # Modes: system (default), light, dark
        # ctk.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green
        
        #=========================university management system=======================================#
   
        self.Frameleft = ctk.CTkFrame(self.master, width=300)
        self.Frameleft.pack(side=LEFT, fill=Y)
        #############################################################################################
        self.Nom = ctk.CTkLabel(self.Frameleft,text='Nom', font=('Helvetica',15))
        self.Nom.place(x=10,y=20 )
        self.Prenom = ctk.CTkLabel(self.Frameleft,text='Prenom', font=('Helvetica',15))
        self.Prenom.place(x=10,y=60 )
        self.Age = ctk.CTkLabel(self.Frameleft,text='Age', font=('Helvetica',15))
        self.Age.place(x=10,y=100)
        self.Motif = ctk.CTkLabel(self.Frameleft,text='Motif', font=('Helvetica',15))
        self.Motif.place(x=10,y=140 )
        self.Jour = ctk.CTkLabel(self.Frameleft,text='Jour', font=('Helvetica',15))
        self.Jour.place(x=10,y=180)
        self.Rendez_vous = ctk.CTkLabel(self.Frameleft,text='Rendez-vous', font=('Helvetica',15))
        self.Rendez_vous.place(x=10,y=220)
        self.Montant_total = ctk.CTkLabel(self.Frameleft,text='Montant total', font=('Helvetica',15))
        self.Montant_total.place(x=10,y=260)
        self.Versement = ctk.CTkLabel(self.Frameleft,text='Versement', font=('Helvetica',15))
        self.Versement.place(x=10,y=300)
        self.Reste = ctk.CTkLabel(self.Frameleft,text='Reste', font=('Helvetica',15))
        self.Reste.place(x=10,y=340)
        self.Tel = ctk.CTkLabel(self.Frameleft,text='Telephone', font=('Helvetica',15))
        self.Tel.place(x=10,y=380)
        
##################################
        
        self.nom = StringVar()
        self.prenom = StringVar()
        self.age = StringVar()
        self.motif = StringVar()
        self.jour = StringVar()
        self.rendez_vous = StringVar()
        self.montant_ttl = StringVar()
        self.versement = StringVar()
        self.reste = StringVar()
        self.num_de_tel = StringVar()

    


    
########################################################
        self.nom_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.nom)
        self.nom_entry.configure(justify="center")
        self.nom_entry.place(x=120,y=20)
        self.prenom_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.prenom)
        self.prenom_entry.configure(justify="center")
        self.prenom_entry.place(x=120,y=60)
        self.age_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.age)
        self.age_entry.configure(justify="center")
        self.age_entry.place(x=120,y=100)
        self.motif_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.motif)
        self.motif_entry.configure(justify="center")
        self.motif_entry.place(x=120,y=140)
        self.jour_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.jour)
        self.jour_entry.configure(justify="center")
        self.jour_entry.place(x=120,y=180)
        self.rendez_vous_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.rendez_vous)
        self.rendez_vous_entry.configure(justify="center")
        self.rendez_vous_entry.place(x=120,y=220)
        self.montant_total_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.montant_ttl)
        self.montant_total_entry.configure(justify="center")
        self.montant_total_entry.place(x=120,y=260)
        self.versement_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable = self.versement)
        self.versement_entry.configure(justify="center")
        self.versement_entry.place(x=120,y=300)
        self.reste_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable =self.reste)
        self.reste_entry.configure(justify="center")
        self.reste_entry.place(x=120,y=340)
        self.tel_entry = ctk.CTkEntry(self.Frameleft, font=('tahoma',12), textvariable =self.num_de_tel)
        self.tel_entry.configure(justify="center")
        self.tel_entry.place(x=120,y=380)


        self.buttonAdd=ctk.CTkButton(self.Frameleft,text='Ajouter', command=self.ajouter,  font=('Helvetica',15,'bold'))
        self.buttonAdd.place(x=10,y=450)

        #fichier = pathlib.Path("C:\\Users\\pc\\assist_cabinet_dentaire\\Liste_patients.xlsx")
        
        ####################################### RIGHT ####################################################
        self.Frameright = ctk.CTkFrame(self.master, height=800, fg_color='White')
        self.Frameright.pack(fill=BOTH, expand=True)
        # ##################################################################################################
        self.Framerighttop = ctk.CTkFrame(self.Frameright, height=70)
         
        self.studentsearch = ctk.CTkEntry(self.Framerighttop,  font=('Helvetica',18,'bold'), width=10)
        self.studentsearch.grid(row = 0, column = 0, sticky='nsew', pady=10, padx=10)
        self.buttonsearch = ctk.CTkButton(self.Framerighttop, text='Rechercher', command=self.export_to_excel, font=('Helvetica',15,'bold'), height=35)
        self.buttonsearch.grid(row = 0, column = 1, sticky='nsew', pady=10, padx=10)
           
        self.Framerighttop.grid_columnconfigure(0, weight=1)
        self.Framerighttop.grid_columnconfigure(0, weight=1)  

        self.Framerighttop.pack(fill=X)

        ##################################################################################################
        



        self.frameView = ctk.CTkFrame(self.Frameright, height=400)
        self.frameView.pack(fill=BOTH)

        self.scrollbar = Scrollbar(self.frameView, orient = VERTICAL)
        

        self.table = ttk.Treeview(self.frameView, column= ("Nom","Prenom","Age","Motif de consultation","Jour","Rendez-vous","Montant total","Versement","Reste","Num de tel"), show='headings', height=17 , yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side=RIGHT, fill=Y)
        self.scrollbar.config(command=self.table.yview())       
        self.table.pack(fill=BOTH)
         


        self.table.heading("Nom",text="Nom")
        self.table.heading("Prenom",text="Prenom")
        self.table.heading("Age",text="Age")
        self.table.heading("Motif de consultation",text="Motif de consultation")
        self.table.heading("Jour",text="Jour")
        self.table.heading("Rendez-vous",text="Rendez-vous")
        self.table.heading("Montant total",text="Montant total")
        self.table.heading("Versement",text="Versement")
        self.table.heading("Reste",text="Reste")
        self.table.heading("Num de tel",text="Num de tel")
       

        self.table.column("Nom", anchor=W, width=5)
        self.table.column("Prenom", anchor=W, width=6)
        self.table.column("Age", anchor=W, width=6)
        self.table.column("Motif de consultation", anchor=W, width=6)
        self.table.column("Jour", anchor=W, width=6)
        self.table.column("Rendez-vous", anchor=W, width=6)
        self.table.column("Montant total", anchor=W, width=6)
        self.table.column("Versement", anchor=W, width=6)
        self.table.column("Reste", anchor=W, width=6)
        self.table.column("Num de tel", anchor=W, width=6)
 
        self.lire()
        self.table.bind("<ButtonRelease>", self.show)


    def ajouter(self):
          nom = self.nom_entry.get()
          prenom = self.prenom_entry.get()
          age = self.age_entry.get()
          motif = self.motif_entry.get()
          jour = self.jour_entry.get()
          rendez_vous = self.rendez_vous_entry.get()
          montant_total = self.montant_total_entry.get()
          versement = self.versement_entry.get()
          reste = self.reste_entry.get()
          tel = self.tel_entry.get()

          # Connect to SQLite database
          conn = sqlite3.connect("data_base.db")
          cursor = conn.cursor()

          # Fetch data from SQLite
          

          if (nom == '' or prenom == '' ) :
           mb.showerror('Erreur','Veuiller saisir le nom et le prenom', parent=self.master)
          else :
              req = "INSERT INTO Patient(Nom, Prenom, Age, Motif, Jour, Rendez_vous, Montant_total, Versement, Reste, Num_de_tel) values ( ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"  
              val = (nom, prenom, age, motif, jour, rendez_vous, montant_total, versement, reste, tel)          
              cursor.execute(req, val)        
              conn.commit()
              conn.close() 
              mb.showinfo('Succes ajoute','Donnees inserees', parent=self.master)
              self.lire()
          
              self.nom_entry.delete(0,'end')
              self.prenom_entry.delete(0,'end')
              self.age_entry.delete(0,'end')
              self.motif_entry.delete(0,'end')
              self.jour_entry.delete(0,'end')
              self.rendez_vous_entry.delete(0,'end')
              self.montant_total_entry.delete(0,'end')
              self.versement_entry.delete(0,'end')
              self.reste_entry.delete(0,'end')
              self.tel_entry.delete(0,'end')

          

          


    def lire(self):
          
          # Connect to SQLite database
          conn = sqlite3.connect("data_base.db")
          cursor = conn.cursor()

          # Fetch data from SQLite
          cursor.execute("SELECT Nom, Prenom, Age, Motif, Jour, Rendez_vous, Montant_total, Versement, Reste, Num_de_tel FROM Patient")
          data = cursor.fetchall()

  
          self.table.delete(*self.table.get_children())

          for i in data:
            self.table.insert('','end', iid=i[0], values=i)

          conn.close()  
              
    def show(self,ev): 
        self.data = self.table.focus()
        alldata = self.table.item(self.data)
        print(self.data)
        val = alldata['values']
        self.nom.set(val[0])
        self.prenom.set(val[1])
        self.age.set(val[2])
        self.motif.set(val[3])
        self.jour.set(val[4])
        self.rendez_vous.set(val[5])
        self.montant_ttl.set(val[6])
        self.versement.set(val[7])
        self.reste.set(val[8])
        self.num_de_tel.set(val[9])

        

    def reset(self):
        self.nom_entry.delete(0,'end')
        self.prenom_entry.delete(0,'end')
        self.age_entry.delete(0,'end')
        self.motif_entry.delete(0,'end')
        self.jour_entry.delete(0,'end')
        self.rendez_vous_entry.delete(0,'end')
        self.montant_total_entry.delete(0,'end')
        self.versement_entry.delete(0,'end')
        self.reste_entry.delete(0,'end')
        self.tel_entry.delete(0,'end')

    def search_value(self):
      
            # Charger le classeur Excel
            wb = load_workbook("C:\\Users\\pc\\assist_cabinet_dentaire\\Liste_patients.xlsx")

            # Sélectionner la feuille de calcul
            sheet = wb.active

            # Valeur à rechercher
            search_value = self.studentsearch.get()

            # Effacer les anciens résultats
            for item in self.table.get_children():
                self.table.delete(item)

            # Rechercher la valeur dans toutes les colonnes
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
                for col_index, cell_value in enumerate(row, start=1):
                    if cell_value == search_value:
                        # Ajouter la ligne complète au Treeview
                        self.table.insert("", "end", values=(row[0], *row[1:]))

            # Si la valeur n'est pas trouvée
            if not self.table.get_children():
                mb.showinfo("Résultat", f"La valeur {search_value} n'a pas été trouvée dans le fichier Excel.")
  
    def export_to_excel(self):
        
            # Connect to SQLite database
            conn = sqlite3.connect("data_base.db")
            cursor = conn.cursor()

            # Fetch data from SQLite
            cursor.execute("SELECT * FROM patient")
            data = cursor.fetchall()

            # Create a new Excel workbook and sheet
            wb = Workbook()
            ws = wb.active

            # Write headers to Excel sheet
            headers = [description[0] for description in cursor.description]
            ws.append(headers)

            # Write data to Excel sheet
            for row in data:
                ws.append(row)

            # Save Excel file
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                wb.save(file_path)
                mb.showinfo("Export Successful", f"Data exported to {file_path}")
            
            conn.close()
        

    
                     




if (__name__ == '__main__'):
    window = ctk.CTk()
    std = Assistante(window)
    mainloop()